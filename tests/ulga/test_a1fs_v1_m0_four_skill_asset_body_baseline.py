from __future__ import annotations

import copy
import json
from io import BytesIO
from pathlib import Path
from zipfile import ZIP_STORED, ZipFile

import pytest
from openpyxl import Workbook

from ulga.builders import build_a1fs_v1_m0_four_skill_asset_body_baseline as builder
from ulga.validators import validate_a1fs_v1_m0_four_skill_asset_body_baseline as validator

SHA = "1" * 40


def _workbook() -> bytes:
    wb = Workbook()
    bodies = wb.active
    bodies.title = "AssetBodies"
    bodies.append(["440 Required Asset Bodies"])
    bodies.append([])
    bodies.append([])
    bodies.append(["asset_id", "lesson_id", "role", "maturity", "disposition"])
    roles = ("CTX", "NTC", "PRD", "CHK", "EVD")
    for lesson in range(88):
        for role in roles:
            bodies.append([
                f"A-{lesson:03d}-{role}", f"L-{lesson:03d}", role,
                "CONTENT_REVIEWED_INTERNAL_D0", "ADMITTED_INTERNAL_D0",
            ])
    bundles = wb.create_sheet("LessonBundles")
    bundles.append(["88 Lesson Bundles"])
    bundles.append([])
    bundles.append([])
    bundles.append(["lesson_id", "bundle_status"])
    for lesson in range(88):
        bundles.append([f"L-{lesson:03d}", "COMPLETE_FIVE_ROLE_BUNDLE"])
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _write_zip(path: Path, rows: dict[str, bytes | str]) -> Path:
    with ZipFile(path, "w", ZIP_STORED) as archive:
        for name, value in rows.items():
            archive.writestr(name, value)
    return path


def _packages(root: Path) -> dict[str, Path]:
    root.mkdir(parents=True, exist_ok=True)
    writing = _write_zip(root / "ketw.zip", {
        "ketw_asset_body/README_INTERNAL_D0.md": "88 lessons; 440 stable required bodies; 0 blockers",
        "ketw_asset_body/A14_INTERNAL_D0_VALIDATOR_AND_EXCEPTION_CLOSEOUT.md": "PASS",
        "ketw_asset_body/A15_INTERNAL_D0_FREEZE_AND_PACKAGING.md": "PASS",
        "outputs/ketw_asset_body/KETW_AB_Asset_Body_Per_Asset_Admission_Ledger_v1.0.0.xlsx": _workbook(),
    })
    result = {"WRITING": writing}
    for skill, spec in builder.MANIFEST_SPECS.items():
        payload = {
            "package_id": spec["package_id"], "version": "1.0.0",
            "status": spec["status"], "gate": spec["gate"],
            **spec["counts"], **spec["extra"],
            "internal_ready": spec["counts"]["lessons"],
            "textual_internal_ready": spec["counts"]["lessons"],
            "learner_release_ready": 0,
        }
        result[skill] = _write_zip(root / f"{skill.casefold()}.zip", {
            spec["manifest"]: json.dumps(payload),
            f"{skill.casefold()}/body.txt": "fixture",
        })
    return result


def _build(paths: dict[str, Path]) -> dict:
    return builder.build_baseline(
        ketw=paths["WRITING"], ketr=paths["READING"],
        kets=paths["SPEAKING"], ketl=paths["LISTENING"],
        source_commit_sha=SHA,
    )


def test_builds_exact_four_skill_frozen_baseline(tmp_path: Path) -> None:
    paths = _packages(tmp_path)
    baseline = _build(paths)
    assert baseline["validation_status"] == builder.STATUS
    assert baseline["totals"] == {"lessons": 414, "asset_bodies": 3320, "teacher_guides": 414}
    assert [row["skill"] for row in baseline["packages"]] == ["WRITING", "READING", "SPEAKING", "LISTENING"]
    assert all(row["release_boundary"]["learner_release_ready"] == 0 for row in baseline["packages"])
    listening = next(row for row in baseline["packages"] if row["skill"] == "LISTENING")
    assert listening["release_boundary"]["audio_deferred"] is True
    assert baseline["claim_boundaries"]["a2_unlocked"] is False


def test_independent_validator_rechecks_hashes_counts_and_boundaries(tmp_path: Path) -> None:
    paths = _packages(tmp_path)
    baseline = _build(paths)
    baseline_path = tmp_path / "baseline.json"
    baseline_path.write_text(json.dumps(baseline), encoding="utf-8")
    report = validator.validate(baseline_path, paths)
    assert report["error_count"] == 0, report["errors"]
    assert report["checked_package_count"] == 4


def test_truncated_writing_archive_fails_closed(tmp_path: Path) -> None:
    paths = _packages(tmp_path)
    original = paths["WRITING"].read_bytes()
    paths["WRITING"].write_bytes(original[: len(original) // 2])
    with pytest.raises(builder.BaselineError, match="zip_unreadable:WRITING"):
        _build(paths)


def test_missing_writing_row_level_ledger_is_rejected(tmp_path: Path) -> None:
    paths = _packages(tmp_path)
    _write_zip(paths["WRITING"], {
        "ketw_asset_body/README_INTERNAL_D0.md": "88 lessons; 440 stable required bodies; 0 blockers",
        "ketw_asset_body/A14_INTERNAL_D0_VALIDATOR_AND_EXCEPTION_CLOSEOUT.md": "PASS",
        "ketw_asset_body/A15_INTERNAL_D0_FREEZE_AND_PACKAGING.md": "PASS",
    })
    with pytest.raises(builder.BaselineError, match="ketw_required_member_missing"):
        _build(paths)


def test_writing_non_admitted_body_is_rejected(tmp_path: Path) -> None:
    paths = _packages(tmp_path)
    data = _workbook()
    workbook = __import__("openpyxl").load_workbook(BytesIO(data))
    workbook["AssetBodies"][5][4].value = "REJECTED"
    stream = BytesIO(); workbook.save(stream)
    _write_zip(paths["WRITING"], {
        "ketw_asset_body/README_INTERNAL_D0.md": "88 lessons; 440 stable required bodies; 0 blockers",
        "ketw_asset_body/A14_INTERNAL_D0_VALIDATOR_AND_EXCEPTION_CLOSEOUT.md": "PASS",
        "ketw_asset_body/A15_INTERNAL_D0_FREEZE_AND_PACKAGING.md": "PASS",
        "outputs/ketw_asset_body/KETW_AB_Asset_Body_Per_Asset_Admission_Ledger_v1.0.0.xlsx": stream.getvalue(),
    })
    with pytest.raises(builder.BaselineError, match="writing_non_admitted_body"):
        _build(paths)


def test_manifest_count_drift_is_rejected(tmp_path: Path) -> None:
    paths = _packages(tmp_path)
    spec = builder.MANIFEST_SPECS["READING"]
    payload = {
        "package_id": spec["package_id"], "version": "1.0.0",
        "status": spec["status"], "gate": spec["gate"],
        **spec["counts"], "asset_bodies": 1009,
        "internal_ready": 101, "learner_release_ready": 0,
    }
    _write_zip(paths["READING"], {spec["manifest"]: json.dumps(payload)})
    with pytest.raises(builder.BaselineError, match="count:READING:asset_bodies"):
        _build(paths)


def test_validator_detects_baseline_hash_tampering(tmp_path: Path) -> None:
    paths = _packages(tmp_path)
    baseline = _build(paths)
    tampered = copy.deepcopy(baseline)
    tampered["packages"][0]["source_sha256"] = "0" * 64
    baseline_path = tmp_path / "tampered.json"
    baseline_path.write_text(json.dumps(tampered), encoding="utf-8")
    report = validator.validate(baseline_path, paths)
    assert report["error_count"] > 0
    assert "hash_mismatch:WRITING" in report["errors"]


def test_invalid_source_commit_sha_is_rejected(tmp_path: Path) -> None:
    paths = _packages(tmp_path)
    with pytest.raises(builder.BaselineError, match="source_commit_sha_invalid"):
        builder.build_baseline(
            ketw=paths["WRITING"], ketr=paths["READING"],
            kets=paths["SPEAKING"], ketl=paths["LISTENING"],
            source_commit_sha="main",
        )
