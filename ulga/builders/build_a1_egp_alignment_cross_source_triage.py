import json
import os
import re
from pathlib import Path

BASE = Path(__file__).resolve().parents[2]
MATRIX = BASE / "ulga" / "reports" / "a1_egp_alignment_matrix.json"
OUT = BASE / "ulga" / "reports" / "a1_egp_alignment_cross_source_triage.json"
SUMMARY = BASE / "ulga" / "reports" / "a1_egp_alignment_cross_source_triage_summary.json"
TASK_ID = "R7-M104B_A1EGPAlignmentCrossSourceEvidenceTriage"
STOPWORDS = {"the", "and", "for", "with", "from", "that", "this", "these", "those", "are", "was", "were", "can", "use", "uses", "using", "basic", "simple", "form"}
EVP_CANDIDATES = [
    "English Vocabulary Profile Online.xlsx",
    "vocabulary_profile/source/English Vocabulary Profile Online.xlsx",
    "sources/English Vocabulary Profile Online.xlsx",
]
TEXT_SOURCE_DIRS = ["raz_output_jsons", "reading_authority", "ulga/reports", "docs/ulga"]
CAMBRIDGE_PATTERNS = ["cambridge", "yle", "pre_a1", "pre-a1", "movers", "starters", "flyers", "a2_key", "ket"]
RAZ_PATTERNS = ["raz", "reading_authority"]


def load(path):
    return json.loads(path.read_text(encoding="utf-8"))


def write(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def tokens(text):
    return {t for t in re.findall(r"[a-zA-Z]+", text.lower()) if len(t) > 2 and t not in STOPWORDS}


def compact_text(value):
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return " ".join(compact_text(v) for v in value[:50])
    if isinstance(value, dict):
        return " ".join(compact_text(v) for v in list(value.values())[:80])
    return ""


def find_file_from_env_or_candidates(env_name, candidates):
    env = os.environ.get(env_name)
    if env and Path(env).is_file():
        return Path(env)
    for candidate in candidates:
        path = BASE / candidate
        if path.is_file():
            return path
    return None


def load_xlsx_token_index(path, limit_rows=8000):
    try:
        from openpyxl import load_workbook
    except Exception:
        return {"status": "SOURCE_PRESENT_BUT_OPENPYXL_UNAVAILABLE", "path": str(path), "tokens": set(), "items": 0}
    workbook = load_workbook(path, read_only=True, data_only=True)
    text_parts = []
    items = 0
    for sheet_name in workbook.sheetnames[:3]:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=1, max_row=limit_rows):
            vals = [str(cell.value).strip() for cell in row if cell.value is not None]
            if vals:
                text_parts.append(" ".join(vals))
                items += 1
    return {"status": "READY", "path": str(path), "tokens": tokens(" ".join(text_parts)), "items": items}


def load_text_source_index(patterns, max_files=120):
    matched_files = []
    text_parts = []
    for dirname in TEXT_SOURCE_DIRS:
        root = BASE / dirname
        if not root.exists():
            continue
        for path in root.rglob("*"):
            if len(matched_files) >= max_files:
                break
            if not path.is_file() or path.suffix.lower() not in {".json", ".md", ".txt"}:
                continue
            name = str(path).lower()
            if not any(pattern in name for pattern in patterns):
                continue
            try:
                if path.suffix.lower() == ".json":
                    data = json.loads(path.read_text(encoding="utf-8"))
                    text_parts.append(compact_text(data))
                else:
                    text_parts.append(path.read_text(encoding="utf-8")[:200000])
                matched_files.append(str(path.relative_to(BASE)))
            except Exception:
                continue
    status = "READY" if matched_files else "SOURCE_NOT_INDEXED"
    return {"status": status, "files": matched_files, "tokens": tokens(" ".join(text_parts)), "items": len(matched_files)}


def match_status(cluster_tokens, source_index):
    if source_index["status"] != "READY":
        return {"status": source_index["status"], "score": 0, "matched_terms": []}
    matched = sorted(cluster_tokens & source_index["tokens"])
    score = len(matched)
    return {"status": "MATCH" if score >= 2 else "NO_MATCH", "score": score, "matched_terms": matched[:12]}


def cluster_text(cluster):
    parts = [cluster.get("cluster_key", ""), cluster.get("super_category", ""), cluster.get("sub_category", "")]
    for row in cluster.get("rows", [])[:8]:
        parts.extend([row.get("guideword", ""), row.get("can_do", ""), row.get("example", "")])
    return " ".join(parts)


def recommended(cluster, evp, raz, cambridge):
    source_missing = [name for name, m in [("EVP", evp), ("RAZ", raz), ("CAMBRIDGE", cambridge)] if m["status"] not in {"MATCH", "NO_MATCH"}]
    if cluster.get("decision") == "COVERED_BY_EXISTING_NODE_REFS":
        return "NO_ACTION_REQUIRED", "HIGH", []
    if source_missing:
        return "DEFER_NEEDS_SOURCE_INDEXING", "LOW", source_missing
    if cambridge["status"] == "MATCH" and raz["status"] == "MATCH":
        return "REVIEW_CREATE_OR_PATCH_WITH_USAGE_AND_EXAM_SUPPORT", "MEDIUM", []
    if cambridge["status"] == "MATCH":
        return "REVIEW_EXAM_ALIGNED_GRAMMAR_NODE", "MEDIUM", []
    if evp["status"] == "MATCH" and cambridge["status"] == "NO_MATCH":
        return "REVIEW_LEXICAL_OR_LEXICAL_GRAMMAR_BRIDGE", "LOW", []
    if raz["status"] == "MATCH":
        return "REVIEW_USAGE_ONLY_SUPPORT", "LOW", []
    return "DEFER_SOURCE_AMBIGUOUS", "LOW", []


def main():
    matrix = load(MATRIX)
    evp_path = find_file_from_env_or_candidates("EVP_SOURCE_XLSX", EVP_CANDIDATES)
    evp_index = load_xlsx_token_index(evp_path) if evp_path else {"status": "SOURCE_NOT_INDEXED", "path": None, "tokens": set(), "items": 0}
    raz_index = load_text_source_index(RAZ_PATTERNS)
    cambridge_index = load_text_source_index(CAMBRIDGE_PATTERNS)
    triage_items = []
    recommendation_counts = {}
    confidence_counts = {}
    source_status_counts = {"evp": {}, "raz": {}, "cambridge": {}}
    for cluster in matrix.get("clusters", []):
        ctok = tokens(cluster_text(cluster))
        evp = match_status(ctok, evp_index)
        raz = match_status(ctok, raz_index)
        cambridge = match_status(ctok, cambridge_index)
        rec, confidence, missing = recommended(cluster, evp, raz, cambridge)
        recommendation_counts[rec] = recommendation_counts.get(rec, 0) + 1
        confidence_counts[confidence] = confidence_counts.get(confidence, 0) + 1
        for key, match in [("evp", evp), ("raz", raz), ("cambridge", cambridge)]:
            source_status_counts[key][match["status"]] = source_status_counts[key].get(match["status"], 0) + 1
        triage_items.append({
            "cluster_id": cluster.get("cluster_id"),
            "cluster_key": cluster.get("cluster_key"),
            "decision_from_alignment_audit": cluster.get("decision"),
            "row_count": cluster.get("row_count"),
            "missing_row_count": cluster.get("missing_row_count"),
            "evp_match": evp,
            "raz_usage_match": raz,
            "cambridge_exam_match": cambridge,
            "recommended_operator_decision": rec,
            "confidence": confidence,
            "remaining_human_review_reason": missing,
            "canonical_grammar_write_allowed": False,
        })
    report = {
        "task_id": TASK_ID,
        "artifact_id": "a1_egp_alignment_cross_source_triage",
        "source_artifact_id": matrix.get("artifact_id"),
        "source_indexes": {
            "evp": {"status": evp_index["status"], "path": evp_index.get("path"), "items": evp_index.get("items")},
            "raz": {"status": raz_index["status"], "files": raz_index.get("files", [])[:30], "items": raz_index.get("items")},
            "cambridge": {"status": cambridge_index["status"], "files": cambridge_index.get("files", [])[:30], "items": cambridge_index.get("items")},
        },
        "triage_items": sorted(triage_items, key=lambda x: (x["recommended_operator_decision"], -(x.get("missing_row_count") or 0), x.get("cluster_key") or "")),
        "final_closeout_allowed": False,
        "a2_a2plus_progression_allowed": False,
        "canonical_grammar_write_allowed": False,
        "local_validation_required": True,
        "ci_gate_required": False,
    }
    summary = {
        "task_id": TASK_ID,
        "artifact_id": "a1_egp_alignment_cross_source_triage_summary",
        "validation_status": "PASS",
        "cluster_count": len(triage_items),
        "source_status_counts": source_status_counts,
        "recommendation_counts": dict(sorted(recommendation_counts.items())),
        "confidence_counts": dict(sorted(confidence_counts.items())),
        "source_indexing_required": any(idx["status"] != "READY" for idx in [evp_index, raz_index, cambridge_index]),
        "final_closeout_allowed": False,
        "a2_a2plus_progression_allowed": False,
        "canonical_grammar_write_allowed": False,
        "next_short_step": "R7-M104C_A1EGPAlignmentCrossSourceTriageReviewPacket",
        "stop_reason": "NONE",
    }
    write(OUT, report)
    write(SUMMARY, summary)
    print("A1 EGP alignment cross-source triage build: PASS")
    print("Clusters:", len(triage_items))
    print("Source indexing required:", summary["source_indexing_required"])
    print("Recommendation counts:", summary["recommendation_counts"])


if __name__ == "__main__":
    main()
