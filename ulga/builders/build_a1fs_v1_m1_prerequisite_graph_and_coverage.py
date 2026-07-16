#!/usr/bin/env python3
"""Build the frozen A1/A1+ four-skill prerequisite graph and coverage registry.

Task: A1FS-V1-M1_A1A1PlusPrerequisiteGraphAndCoverage
This stage defines the mastery denominator and the A2 lock contract.  It does
not implement learner state, unlock A2, release private content, or claim that
any learner has mastered a node.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
from collections import defaultdict
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

TASK_ID = "A1FS-V1-M1_A1A1PlusPrerequisiteGraphAndCoverage"
SCHEMA_VERSION = "a1fs.v1.m1.prerequisite_graph_and_coverage.v1"
STATUS = "PASS_A1FS_V1_M1_PREREQUISITE_GRAPH_AND_COVERAGE"
NEXT_SHORT_STEP = "A1FS-V1-M2_FourSkillAssetBodyConsumerAndQuery"
BASELINE_STATUS = "PASS_A1FS_V1_M0_FOUR_SKILL_ASSET_BODY_BASELINE_FROZEN"
LEVEL_RANK = {"A1": 0, "A1+": 1, "A2": 2}


class GraphError(ValueError):
    """Fail-closed graph construction error."""


@dataclass
class RefUse:
    levels: set[str] = field(default_factory=set)
    lesson_ids: set[str] = field(default_factory=set)
    asset_body_ids: set[str] = field(default_factory=set)
    roles: set[str] = field(default_factory=set)
    mastery: bool = False


@dataclass
class Lesson:
    skill: str
    lesson_id: str
    level: str
    order_key: tuple[Any, ...]
    assets: set[str] = field(default_factory=set)
    roles: set[str] = field(default_factory=set)
    refs: dict[str, bool] = field(default_factory=dict)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _atomic_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.replace(temporary, path)


def _read_json_member(archive: ZipFile, name: str) -> dict[str, Any]:
    try:
        value = json.loads(archive.read(name).decode("utf-8"))
    except (KeyError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise GraphError(f"json_member_unreadable:{name}:{exc}") from exc
    if not isinstance(value, dict):
        raise GraphError(f"json_member_not_object:{name}")
    return value


def _check_archive(path: Path, skill: str) -> ZipFile:
    if not path.is_file():
        raise GraphError(f"package_missing:{skill}:{path}")
    try:
        archive = ZipFile(path)
        bad = archive.testzip()
    except (OSError, BadZipFile) as exc:
        raise GraphError(f"zip_unreadable:{skill}:{exc}") from exc
    if bad is not None:
        archive.close()
        raise GraphError(f"zip_crc_failure:{skill}:{bad}")
    return archive


def _level_for(skill: str, lesson_id: str, explicit: Any = None) -> str:
    if explicit in LEVEL_RANK:
        return str(explicit)
    markers = {
        "LISTENING": (("-LF-", "A1"), ("-LB-", "A1+"), ("-KL-", "A2")),
        "READING": (("-RF-", "A1"), ("-RB-", "A1+"), ("-KR-", "A2")),
    }
    for marker, level in markers.get(skill, ()):
        if marker in lesson_id:
            return level
    raise GraphError(f"lesson_level_unresolved:{skill}:{lesson_id}:{explicit}")


def _natural_key(value: str) -> tuple[Any, ...]:
    return tuple(int(part) if part.isdigit() else part for part in re.split(r"(\d+)", value))


def _lesson_map_from_json(path: Path, skill: str, member: str) -> dict[str, Lesson]:
    with _check_archive(path, skill) as archive:
        payload = _read_json_member(archive, member)
    rows = payload.get("rows") or payload.get("asset_bodies")
    if not isinstance(rows, list) or not rows:
        raise GraphError(f"asset_registry_empty:{skill}")
    lessons: dict[str, Lesson] = {}
    for row in rows:
        if not isinstance(row, dict):
            raise GraphError(f"asset_row_not_object:{skill}")
        lesson_id = str(row.get("lesson_id") or "")
        asset_id = str(row.get("asset_body_id") or "")
        role = str(row.get("role") or "")
        if not lesson_id or not asset_id or not role:
            raise GraphError(f"asset_identity_missing:{skill}")
        level = _level_for(skill, lesson_id, row.get("level"))
        lesson = lessons.setdefault(
            lesson_id,
            Lesson(skill, lesson_id, level, (LEVEL_RANK[level], _natural_key(lesson_id))),
        )
        if lesson.level != level or asset_id in lesson.assets:
            raise GraphError(f"asset_or_level_collision:{skill}:{asset_id}")
        lesson.assets.add(asset_id)
        lesson.roles.add(role)
        field_policy: tuple[tuple[str, bool], ...]
        if skill == "LISTENING":
            field_policy = (("target_refs", True), ("candidate_resource_refs", False))
        elif skill == "READING":
            field_policy = (("target_refs", True), ("supporting_target_refs", False))
        else:
            field_policy = (("resource_refs", True),)
        for field_name, default_mastery in field_policy:
            refs = row.get(field_name) or []
            if isinstance(refs, str):
                refs = [refs]
            if not isinstance(refs, list):
                raise GraphError(f"reference_field_invalid:{skill}:{asset_id}:{field_name}")
            for ref in refs:
                ref = str(ref).strip()
                if not ref:
                    continue
                mastery = default_mastery
                if skill == "SPEAKING" and ref.startswith(("CHK-", "SIT-")):
                    mastery = False
                lesson.refs[ref] = lesson.refs.get(ref, False) or mastery
    return lessons


def _sheet_records(workbook_bytes: bytes, sheet: str, id_header: str) -> list[dict[str, Any]]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        raise GraphError(f"writing_sheet_missing:{sheet}")
    rows = list(workbook[sheet].iter_rows(values_only=True))
    index = next((i for i, row in enumerate(rows) if row and row[0] == id_header), None)
    if index is None:
        raise GraphError(f"writing_header_missing:{sheet}:{id_header}")
    headers = [str(value) if value is not None else "" for value in rows[index]]
    return [
        {headers[i]: value for i, value in enumerate(row) if i < len(headers) and headers[i]}
        for row in rows[index + 1 :]
        if row and row[0] is not None
    ]


def _expand_writing_targets(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    if text.casefold().startswith("all "):
        return [text]
    expanded: list[str] = []
    current_prefix: str | None = None
    for clause in re.split(r"\s*;\s*", text):
        clause = re.sub(r"\s+prep$", "", clause.strip())
        for token in clause.split("/"):
            token = token.strip()
            match = re.fullmatch(r"([A-Z0-9]+-)(\d+)(?:[–-](\d+))?", token)
            if match:
                current_prefix = match.group(1)
                start = int(match.group(2)); end = int(match.group(3) or start)
                width = len(match.group(2))
            elif current_prefix and re.fullmatch(r"\d+(?:[–-]\d+)?", token):
                parts = re.split(r"[–-]", token)
                start = int(parts[0]); end = int(parts[-1]); width = len(parts[0])
            else:
                raise GraphError(f"writing_target_unparseable:{value}:{token}")
            if end < start or end - start > 100:
                raise GraphError(f"writing_target_range_invalid:{value}")
            expanded.extend(f"{current_prefix}{number:0{width}d}" for number in range(start, end + 1))
    return sorted(set(expanded), key=_natural_key)


def _writing_lessons(path: Path) -> dict[str, Lesson]:
    member = "outputs/ketw_asset_body/KETW_AB_Asset_Body_Per_Asset_Admission_Ledger_v1.0.0.xlsx"
    with _check_archive(path, "WRITING") as archive:
        try:
            workbook_bytes = archive.read(member)
        except KeyError as exc:
            raise GraphError(f"writing_ledger_missing:{member}") from exc
    rows = _sheet_records(workbook_bytes, "AssetBodies", "asset_id")
    if not rows:
        raise GraphError("asset_registry_empty:WRITING")
    lessons: dict[str, Lesson] = {}
    stage_level = {"A1 WF": "A1", "A1+ WB": "A1+", "A2 KP": "A2"}
    for row in rows:
        lesson_id = str(row.get("lesson_id") or "")
        asset_id = str(row.get("asset_id") or "")
        role = str(row.get("role") or "")
        stage = str(row.get("stage") or "")
        if not lesson_id or not asset_id or not role or stage not in stage_level:
            raise GraphError(f"writing_asset_identity_or_stage_invalid:{asset_id}")
        level = stage_level[stage]
        unit = str(row.get("unit") or "")
        sequence = int(row.get("sequence") or 0)
        lesson = lessons.setdefault(
            lesson_id,
            Lesson("WRITING", lesson_id, level, (LEVEL_RANK[level], _natural_key(unit), sequence, _natural_key(lesson_id))),
        )
        if lesson.level != level or asset_id in lesson.assets:
            raise GraphError(f"asset_or_level_collision:WRITING:{asset_id}")
        lesson.assets.add(asset_id); lesson.roles.add(role)
        for ref in _expand_writing_targets(row.get("capability_targets")):
            lesson.refs[ref] = True
    return lessons


def _baseline(path: Path, package_paths: dict[str, Path]) -> tuple[dict[str, Any], str]:
    try:
        raw = path.read_bytes(); value = json.loads(raw)
    except (OSError, json.JSONDecodeError) as exc:
        raise GraphError(f"baseline_unreadable:{exc}") from exc
    if not isinstance(value, dict) or value.get("validation_status") != BASELINE_STATUS:
        raise GraphError("baseline_status_invalid")
    packages = value.get("packages")
    if not isinstance(packages, list) or len(packages) != 4:
        raise GraphError("baseline_package_count_invalid")
    expected = {row.get("skill"): row.get("source_sha256") for row in packages if isinstance(row, dict)}
    for skill, package_path in package_paths.items():
        actual = _sha256(package_path)
        if expected.get(skill) != actual:
            raise GraphError(f"baseline_package_hash_mismatch:{skill}")
    return value, hashlib.sha256(raw).hexdigest()


def _acyclic(node_ids: set[str], edges: Iterable[dict[str, str]]) -> bool:
    incoming = {node: 0 for node in node_ids}; outgoing: dict[str, list[str]] = defaultdict(list)
    for edge in edges:
        source = edge["from_node_id"]; target = edge["to_node_id"]
        if source not in node_ids or target not in node_ids:
            return False
        outgoing[source].append(target); incoming[target] += 1
    queue = [node for node, count in incoming.items() if count == 0]
    visited = 0
    while queue:
        node = queue.pop(); visited += 1
        for target in outgoing[node]:
            incoming[target] -= 1
            if incoming[target] == 0:
                queue.append(target)
    return visited == len(node_ids)


def build_graph(*, baseline_path: Path, ketw: Path, ketr: Path, kets: Path, ketl: Path) -> dict[str, Any]:
    paths = {"WRITING": ketw, "READING": ketr, "SPEAKING": kets, "LISTENING": ketl}
    baseline, baseline_sha = _baseline(baseline_path, paths)
    by_skill = {
        "WRITING": _writing_lessons(ketw),
        "READING": _lesson_map_from_json(ketr, "READING", "ketr_asset_body_production/data/asset_body_registry.json"),
        "SPEAKING": _lesson_map_from_json(kets, "SPEAKING", "07_KETS_AB/data/asset_body_registry.json"),
        "LISTENING": _lesson_map_from_json(ketl, "LISTENING", "ketl_asset_body_production/data/asset_body_registry.json"),
    }
    expected_counts = {row["skill"]: row["counts"] for row in baseline["packages"]}
    for skill, lessons in by_skill.items():
        asset_count = sum(len(lesson.assets) for lesson in lessons.values())
        if len(lessons) != int(expected_counts[skill]["lessons"]):
            raise GraphError(f"baseline_lesson_count_mismatch:{skill}")
        if asset_count != int(expected_counts[skill]["asset_bodies"]):
            raise GraphError(f"baseline_asset_count_mismatch:{skill}")

    nodes: list[dict[str, Any]] = []
    edges: list[dict[str, str]] = []
    coverage: list[dict[str, Any]] = []
    ref_uses: dict[tuple[str, str], RefUse] = defaultdict(RefUse)
    a2_lessons: list[str] = []
    required_lesson_nodes: list[str] = []

    for skill in ("LISTENING", "SPEAKING", "READING", "WRITING"):
        ordered = sorted(by_skill[skill].values(), key=lambda row: row.order_key)
        required = [row for row in ordered if row.level in {"A1", "A1+"}]
        handoff = [row for row in ordered if row.level == "A2"]
        if not required or not handoff:
            raise GraphError(f"level_partition_incomplete:{skill}")
        previous: str | None = None
        for lesson in ordered:
            node_id = f"LESSON:{skill}:{lesson.lesson_id}"
            is_required = lesson.level in {"A1", "A1+"}
            nodes.append({
                "node_id": node_id, "node_type": "LESSON", "skill": skill,
                "level": lesson.level, "source_ref": lesson.lesson_id,
                "mastery_required_before_a2": is_required,
                "asset_body_count": len(lesson.assets), "roles": sorted(lesson.roles),
            })
            if is_required:
                required_lesson_nodes.append(node_id)
            else:
                a2_lessons.append(node_id)
            if previous is not None and (is_required or lesson.level == "A2"):
                edges.append({"from_node_id": previous, "to_node_id": node_id, "edge_type": "PRECEDES"})
            previous = node_id
            for ref, mastery in lesson.refs.items():
                use = ref_uses[(skill, ref)]
                use.levels.add(lesson.level); use.lesson_ids.add(lesson.lesson_id)
                use.asset_body_ids.update(lesson.assets); use.roles.update(lesson.roles)
                use.mastery = use.mastery or mastery

    reference_node_ids: dict[tuple[str, str], str] = {}
    required_reference_nodes: list[str] = []
    for (skill, ref), use in sorted(ref_uses.items()):
        level = min(use.levels, key=lambda item: LEVEL_RANK[item])
        node_id = f"REF:{skill}:{ref}"
        reference_node_ids[(skill, ref)] = node_id
        required = use.mastery and level in {"A1", "A1+"}
        nodes.append({
            "node_id": node_id, "node_type": "CAPABILITY" if use.mastery else "SUPPORT_RESOURCE",
            "skill": skill, "level": level, "source_ref": ref,
            "mastery_required_before_a2": required,
        })
        if required:
            required_reference_nodes.append(node_id)
        coverage.append({
            "node_id": node_id, "skill": skill, "source_ref": ref,
            "coverage_class": "MASTERY" if use.mastery else "SUPPORT",
            "levels": sorted(use.levels, key=lambda item: LEVEL_RANK[item]),
            "lesson_ids": sorted(use.lesson_ids, key=_natural_key),
            "asset_body_ids": sorted(use.asset_body_ids, key=_natural_key),
            "roles": sorted(use.roles), "coverage_status": "COVERED",
        })
        for lesson_id in use.lesson_ids:
            edges.append({
                "from_node_id": node_id,
                "to_node_id": f"LESSON:{skill}:{lesson_id}",
                "edge_type": "TAUGHT_BY",
            })

    gate_id = "GATE:A1FS:A2_LOCK"
    nodes.append({
        "node_id": gate_id, "node_type": "A2_LOCK", "skill": "FOUR_SKILL",
        "level": "A2", "source_ref": "A2_ENTRY",
        "mastery_required_before_a2": False,
    })
    required_mastery = sorted(set(required_lesson_nodes + required_reference_nodes))
    for node_id in required_mastery:
        edges.append({"from_node_id": node_id, "to_node_id": gate_id, "edge_type": "UNLOCK_REQUIRES"})
    # Remove direct A1+ -> A2 sequencing: every A2 track is gated by the global lock.
    required_set = set(required_lesson_nodes)
    a2_set = set(a2_lessons)
    edges = [edge for edge in edges if not (
        edge["edge_type"] == "PRECEDES" and edge["from_node_id"] in required_set and edge["to_node_id"] in a2_set
    )]
    first_a2_by_skill: dict[str, str] = {}
    for node_id in a2_lessons:
        skill = node_id.split(":", 2)[1]
        first_a2_by_skill.setdefault(skill, node_id)
    for node_id in first_a2_by_skill.values():
        edges.append({"from_node_id": gate_id, "to_node_id": node_id, "edge_type": "LOCKS"})

    node_ids = {node["node_id"] for node in nodes}
    if len(node_ids) != len(nodes):
        raise GraphError("duplicate_node_id")
    edge_keys = {(e["from_node_id"], e["to_node_id"], e["edge_type"]) for e in edges}
    edges = [dict(zip(("from_node_id", "to_node_id", "edge_type"), key)) for key in sorted(edge_keys)]
    if not _acyclic(node_ids, edges):
        raise GraphError("graph_not_acyclic_or_edge_unresolved")
    if any(not row["lesson_ids"] or not row["asset_body_ids"] for row in coverage):
        raise GraphError("coverage_orphan")

    level_counts = {level: 0 for level in LEVEL_RANK}
    for node in nodes:
        if node["node_type"] == "LESSON":
            level_counts[node["level"]] += 1
    return {
        "task_id": TASK_ID, "schema_version": SCHEMA_VERSION, "validation_status": STATUS,
        "source_baseline_sha256": baseline_sha,
        "nodes": sorted(nodes, key=lambda row: row["node_id"]),
        "edges": edges, "coverage": sorted(coverage, key=lambda row: row["node_id"]),
        "counts": {
            "node_count": len(nodes), "edge_count": len(edges), "coverage_record_count": len(coverage),
            "lesson_count": sum(level_counts.values()), "lesson_count_by_level": level_counts,
            "required_mastery_node_count": len(required_mastery),
            "a2_handoff_lesson_count": len(a2_lessons), "uncovered_required_node_count": 0,
        },
        "a2_lock_contract": {
            "gate_node_id": gate_id, "state": "LOCKED_BY_DESIGN",
            "required_mastery_node_ids": required_mastery,
            "a2_handoff_lesson_node_ids": sorted(a2_lessons),
            "unlock_rule": "ALL_REQUIRED_MASTERY_NODES_MUST_BE_MASTERED",
            "runtime_unlock_implemented": False,
        },
        "claim_boundaries": {
            "source_packages_committed": False, "asset_body_content_modified": False,
            "learner_release_approved": False, "mastery_claimed": False,
            "a2_unlocked": False, "runtime_planner_implemented": False,
            "human_pilot_claimed": False, "listening_audio_complete": False,
        },
        "errors": [], "next_short_step": NEXT_SHORT_STEP,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--baseline", type=Path, required=True)
    parser.add_argument("--ketw", type=Path, required=True)
    parser.add_argument("--ketr", type=Path, required=True)
    parser.add_argument("--kets", type=Path, required=True)
    parser.add_argument("--ketl", type=Path, required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    args = parser.parse_args()
    graph = build_graph(
        baseline_path=args.baseline, ketw=args.ketw, ketr=args.ketr,
        kets=args.kets, ketl=args.ketl,
    )
    output = args.output_root / "a1a1plus_prerequisite_graph_and_coverage.private.json"
    _atomic_json(output, graph)
    print(json.dumps({
        "validation_status": graph["validation_status"], **graph["counts"],
        "a2_lock_state": graph["a2_lock_contract"]["state"],
        "next_short_step": graph["next_short_step"], "output": str(output),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
