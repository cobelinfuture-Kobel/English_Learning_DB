import argparse
import json
import os
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parents[2]
OUT_PATH = BASE_DIR / "ulga" / "reports" / "egp_row_index_compact.json"
SUMMARY_PATH = BASE_DIR / "ulga" / "reports" / "egp_row_index_compact_summary.json"
TASK_ID = "R7-M97B_EGPCompactRowIndexBuilder"
DEFAULT_SOURCE_NAMES = [
    "English Grammar Profile Online.xlsx",
    "data/English Grammar Profile Online.xlsx",
    "sources/English Grammar Profile Online.xlsx",
]


def write_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def find_source(explicit_source=None):
    candidates = []
    if explicit_source:
        candidates.append(Path(explicit_source))
    env_source = os.environ.get("EGP_SOURCE_XLSX")
    if env_source:
        candidates.append(Path(env_source))
    candidates.extend(BASE_DIR / name for name in DEFAULT_SOURCE_NAMES)
    for path in candidates:
        if path.exists() and path.is_file():
            return path
    return None


def cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def build_from_workbook(source_path):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required to build the EGP compact row index") from exc
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet_name = "Data" if "Data" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    headers = [cell_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [cell_text(cell.value) for cell in row]
        if not any(values):
            continue
        row_id = values[0] if values else ""
        source_ref = f"EGP_SOURCE_XLSX::{sheet_name}!A{row_number}:H{row_number}::id={row_id}"
        rows.append({
            "row_number": row_number,
            "source_ref": source_ref,
            "row_id": row_id,
            "level": values[1] if len(values) > 1 else "",
            "super_category": values[2] if len(values) > 2 else "",
            "sub_category": values[3] if len(values) > 3 else "",
            "guideword": values[4] if len(values) > 4 else "",
            "can_do": values[5] if len(values) > 5 else "",
        })
    return sheet_name, headers, rows


def build_missing_source_report():
    report = {
        "task_id": TASK_ID,
        "artifact_id": "egp_row_index_compact",
        "source_workbook_status": "MISSING",
        "rows": [],
        "scope_constraints": {
            "canonical_grammar_write_allowed": False,
            "coverage_increase_allowed": False,
            "runtime_change_allowed": False,
        },
    }
    summary = {
        "task_id": TASK_ID,
        "artifact_id": "egp_row_index_compact_summary",
        "validation_status": "PASS_WITH_SOURCE_WORKBOOK_REQUIRED",
        "source_workbook_status": "MISSING",
        "row_count": 0,
        "canonical_grammar_write_allowed": False,
        "next_short_step": "R7-M97B_LocalEGPCompactRowIndexBuild",
        "stop_reason": "SOURCE_WORKBOOK_REQUIRED",
    }
    return report, summary


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=None)
    parser.add_argument("--require-source", action="store_true")
    args = parser.parse_args(argv)
    source_path = find_source(args.source)
    if source_path is None:
        report, summary = build_missing_source_report()
        write_json(OUT_PATH, report)
        write_json(SUMMARY_PATH, summary)
        print("EGP compact row index build: SOURCE_WORKBOOK_REQUIRED")
        if args.require_source:
            return 2
        return 0
    sheet_name, headers, rows = build_from_workbook(source_path)
    report = {
        "task_id": TASK_ID,
        "artifact_id": "egp_row_index_compact",
        "source_workbook_status": "READY",
        "source_workbook_name": source_path.name,
        "sheet_name": sheet_name,
        "headers": headers[:8],
        "rows": rows,
        "scope_constraints": {
            "canonical_grammar_write_allowed": False,
            "coverage_increase_allowed": False,
            "runtime_change_allowed": False,
        },
    }
    summary = {
        "task_id": TASK_ID,
        "artifact_id": "egp_row_index_compact_summary",
        "validation_status": "PASS",
        "source_workbook_status": "READY",
        "source_workbook_name": source_path.name,
        "sheet_name": sheet_name,
        "row_count": len(rows),
        "canonical_grammar_write_allowed": False,
        "next_short_step": "R7-M97C_A1A1PLUSBulkEGPRowCandidateResolverWithCompactIndex",
        "stop_reason": "NONE",
    }
    write_json(OUT_PATH, report)
    write_json(SUMMARY_PATH, summary)
    print("EGP compact row index build: PASS")
    print("Rows:", len(rows))
    return 0


if __name__ == "__main__":
    sys.exit(main())
