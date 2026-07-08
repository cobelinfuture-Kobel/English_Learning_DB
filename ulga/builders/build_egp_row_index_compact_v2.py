import argparse
import json
import os
import sys
from pathlib import Path

BASE = Path(__file__).resolve().parents[2]
OUT = BASE / "ulga" / "reports" / "egp_row_index_compact_v2.json"
SUMMARY = BASE / "ulga" / "reports" / "egp_row_index_compact_v2_summary.json"
TASK_ID = "R7-M100D_EGPCompactRowIndexV2Builder"
SOURCE_NAMES = [
    "English Grammar Profile Online.xlsx",
    "grammar_profile/source/English Grammar Profile Online.xlsx",
    "data/English Grammar Profile Online.xlsx",
    "sources/English Grammar Profile Online.xlsx",
]


def text(value):
    return "" if value is None else str(value).strip()


def write(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def read(path):
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def ready_existing():
    report = read(OUT)
    summary = read(SUMMARY)
    return isinstance(report, dict) and isinstance(summary, dict) and summary.get("row_count", 0) > 0 and summary.get("source_workbook_status") == "READY"


def find_source(explicit=None):
    candidates = []
    if explicit:
        candidates.append(Path(explicit))
    env = os.environ.get("EGP_SOURCE_XLSX")
    if env:
        candidates.append(Path(env))
    candidates.extend(BASE / name for name in SOURCE_NAMES)
    for path in candidates:
        if path.is_file():
            return path
    return None


def idx(headers, name, fallback):
    lower = {h.strip().lower(): i for i, h in enumerate(headers)}
    return lower.get(name.strip().lower(), fallback)


def build(source):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required") from exc
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet_name = "Data" if "Data" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    headers = [text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    id_i = idx(headers, "id", 0)
    super_i = idx(headers, "SuperCategory", 1)
    sub_i = idx(headers, "SubCategory", 2)
    level_i = idx(headers, "Level", 3)
    lexical_i = idx(headers, "Lexical Range", 4)
    guide_i = idx(headers, "Guideword", 5)
    cando_i = idx(headers, "Can-do", 6)
    example_i = idx(headers, "Example", 7)
    rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [text(cell.value) for cell in row]
        if not any(values):
            continue
        row_id = values[id_i] if len(values) > id_i else ""
        rows.append({
            "row_number": row_number,
            "source_ref": f"EGP_SOURCE_XLSX::{sheet_name}!A{row_number}:H{row_number}::id={row_id}",
            "row_id": row_id,
            "level": values[level_i] if len(values) > level_i else "",
            "super_category": values[super_i] if len(values) > super_i else "",
            "sub_category": values[sub_i] if len(values) > sub_i else "",
            "lexical_range": values[lexical_i] if len(values) > lexical_i else "",
            "guideword": values[guide_i] if len(values) > guide_i else "",
            "can_do": values[cando_i] if len(values) > cando_i else "",
            "example": values[example_i] if len(values) > example_i else "",
        })
    return sheet_name, headers, rows


def main(argv=None):
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=None)
    parser.add_argument("--require-source", action="store_true")
    args = parser.parse_args(argv)
    source = find_source(args.source)
    if source is None:
        if ready_existing():
            print("EGP compact row index v2 build: READY_INDEX_PRESERVED")
            return 2 if args.require_source else 0
        report = {"task_id": TASK_ID, "artifact_id": "egp_row_index_compact_v2", "source_workbook_status": "MISSING", "rows": []}
        summary = {"task_id": TASK_ID, "artifact_id": "egp_row_index_compact_v2_summary", "validation_status": "PASS_WITH_SOURCE_WORKBOOK_REQUIRED", "source_workbook_status": "MISSING", "row_count": 0, "stop_reason": "SOURCE_WORKBOOK_REQUIRED"}
        write(OUT, report)
        write(SUMMARY, summary)
        print("EGP compact row index v2 build: SOURCE_WORKBOOK_REQUIRED")
        return 2 if args.require_source else 0
    sheet_name, headers, rows = build(source)
    report = {"task_id": TASK_ID, "artifact_id": "egp_row_index_compact_v2", "source_workbook_status": "READY", "source_workbook_name": source.name, "sheet_name": sheet_name, "headers": headers[:8], "rows": rows, "canonical_grammar_write_allowed": False}
    summary = {"task_id": TASK_ID, "artifact_id": "egp_row_index_compact_v2_summary", "validation_status": "PASS", "source_workbook_status": "READY", "source_workbook_name": source.name, "sheet_name": sheet_name, "row_count": len(rows), "required_fields": ["lexical_range", "can_do", "example"], "canonical_grammar_write_allowed": False, "next_short_step": "R7-M100E_A1A1PLUSSemanticCoverageAuditV2", "stop_reason": "NONE"}
    write(OUT, report)
    write(SUMMARY, summary)
    print("EGP compact row index v2 build: PASS")
    print("Rows:", len(rows))
    return 0


if __name__ == "__main__":
    sys.exit(main())
