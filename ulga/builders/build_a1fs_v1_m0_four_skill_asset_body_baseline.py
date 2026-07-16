#!/usr/bin/env python3
"""Freeze the four private A1FS-V1 Asset Body packages as a metadata baseline.

Task: A1FS-V1-M0_FourSkillTerminalScopeAndFrozenAssetBodyBaseline
Scope: package integrity, frozen identities/counts, release boundaries, hashes.
Allowed outputs: metadata-only JSON below the caller-selected output root.
Forbidden: package extraction, learner release, authority promotion, A2 unlock.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

TASK_ID = "A1FS-V1-M0_FourSkillTerminalScopeAndFrozenAssetBodyBaseline"
SCHEMA_VERSION = "a1fs.v1.m0.four_skill_asset_body_baseline.v1"
STATUS = "PASS_A1FS_V1_M0_FOUR_SKILL_ASSET_BODY_BASELINE_FROZEN"
NEXT_SHORT_STEP = "A1FS-V1-M1_A1A1PlusPrerequisiteGraphAndCoverage"


class BaselineError(ValueError):
    """Fail-closed baseline error."""


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _atomic_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.replace(temporary, path)


def _safe_members(archive: ZipFile, label: str) -> list[str]:
    names = archive.namelist()
    if not names:
        raise BaselineError(f"empty_archive:{label}")
    if len(names) != len(set(names)):
        raise BaselineError(f"duplicate_zip_member:{label}")
    for name in names:
        path = PurePosixPath(name)
        if path.is_absolute() or ".." in path.parts or "\\" in name:
            raise BaselineError(f"unsafe_zip_member:{label}:{name}")
        if name.startswith("/"):
            raise BaselineError(f"unsafe_zip_member:{label}:{name}")
    bad = archive.testzip()
    if bad is not None:
        raise BaselineError(f"zip_crc_failure:{label}:{bad}")
    return names


def _read_json(archive: ZipFile, name: str, label: str) -> dict[str, Any]:
    try:
        value = json.loads(archive.read(name).decode("utf-8"))
    except (KeyError, UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise BaselineError(f"manifest_unreadable:{label}:{name}:{exc}") from exc
    if not isinstance(value, dict):
        raise BaselineError(f"manifest_not_object:{label}:{name}")
    return value


def _require(actual: Any, expected: Any, code: str) -> None:
    if actual != expected:
        raise BaselineError(f"{code}:expected={expected!r}:actual={actual!r}")


MANIFEST_SPECS: dict[str, dict[str, Any]] = {
    "LISTENING": {
        "manifest": "ketl_asset_body_production/manifest.json",
        "package_id": "English_Grammar_A1_KETL_Asset_Body_Teacher_Delivery_Internal_D0_AudioDeferred_v1.0.0",
        "status": "FROZEN_INTERNAL_D0_WITH_AUDIO_DEFERRED",
        "gate": "PASS_ACCEPTED_INTERNAL_D0_WITH_AUDIO_DEFERRED",
        "counts": {"lessons": 130, "asset_bodies": 1300, "teacher_guides": 130},
        "extra": {"original_audio_scripts": 260, "audio_assets": 0},
    },
    "READING": {
        "manifest": "ketr_asset_body_production/manifest.json",
        "package_id": "English_Grammar_A1_KETR_Asset_Body_Teacher_Delivery_Internal_D0_v1.0.0",
        "status": "FROZEN_INTERNAL_D0",
        "gate": "PASS_ACCEPTED_INTERNAL_D0",
        "counts": {"lessons": 101, "asset_bodies": 1010, "teacher_guides": 101},
        "extra": {},
    },
    "SPEAKING": {
        "manifest": "07_KETS_AB/manifest.json",
        "package_id": "English_Grammar_A1_KETS_Asset_Body_Teacher_Delivery_Internal_D0_v1.0.0",
        "status": "FROZEN_INTERNAL_D0",
        "gate": "PASS_ACCEPTED_INTERNAL_D0",
        "counts": {"lessons": 95, "asset_bodies": 570, "teacher_guides": 95},
        "extra": {"audio_assets": 0},
    },
}


def _manifest_package(path: Path, skill: str) -> dict[str, Any]:
    spec = MANIFEST_SPECS[skill]
    try:
        with ZipFile(path) as archive:
            names = _safe_members(archive, skill)
            manifest = _read_json(archive, spec["manifest"], skill)
    except (OSError, BadZipFile) as exc:
        raise BaselineError(f"zip_unreadable:{skill}:{path}:{exc}") from exc
    _require(manifest.get("package_id"), spec["package_id"], f"package_id:{skill}")
    _require(manifest.get("status"), spec["status"], f"status:{skill}")
    _require(manifest.get("gate"), spec["gate"], f"gate:{skill}")
    for key, expected in {**spec["counts"], **spec["extra"]}.items():
        _require(manifest.get(key), expected, f"count:{skill}:{key}")
    return {
        "skill": skill,
        "source_filename": path.name,
        "source_sha256": _sha256(path),
        "zip_entry_count": len(names),
        "package_id": manifest["package_id"],
        "version": manifest.get("version"),
        "status": manifest["status"],
        "gate": manifest["gate"],
        "counts": dict(spec["counts"]),
        "release_boundary": {
            "internal_ready": int(manifest.get("internal_ready", manifest.get("textual_internal_ready", 0))),
            "learner_release_ready": int(manifest.get("learner_release_ready", 0)),
            "audio_deferred": skill == "LISTENING",
        },
    }


def _sheet_rows(workbook_bytes: bytes, sheet: str, id_header: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    if sheet not in workbook.sheetnames:
        raise BaselineError(f"ketw_sheet_missing:{sheet}")
    rows = list(workbook[sheet].iter_rows(values_only=True))
    header_index = next((i for i, row in enumerate(rows) if row and row[0] == id_header), None)
    if header_index is None:
        raise BaselineError(f"ketw_header_missing:{sheet}:{id_header}")
    headers = [str(value) if value is not None else "" for value in rows[header_index]]
    values = [
        {headers[i]: value for i, value in enumerate(row) if i < len(headers) and headers[i]}
        for row in rows[header_index + 1 :]
        if row and row[0] is not None
    ]
    return headers, values


def _writing_package(path: Path) -> dict[str, Any]:
    workbook_name = "outputs/ketw_asset_body/KETW_AB_Asset_Body_Per_Asset_Admission_Ledger_v1.0.0.xlsx"
    required = {
        "ketw_asset_body/README_INTERNAL_D0.md",
        "ketw_asset_body/A14_INTERNAL_D0_VALIDATOR_AND_EXCEPTION_CLOSEOUT.md",
        "ketw_asset_body/A15_INTERNAL_D0_FREEZE_AND_PACKAGING.md",
        workbook_name,
    }
    try:
        with ZipFile(path) as archive:
            names = _safe_members(archive, "WRITING")
            missing = sorted(required - set(names))
            if missing:
                raise BaselineError(f"ketw_required_member_missing:{missing}")
            readme = archive.read("ketw_asset_body/README_INTERNAL_D0.md").decode("utf-8")
            body_bytes = archive.read(workbook_name)
    except (OSError, BadZipFile, UnicodeDecodeError) as exc:
        raise BaselineError(f"zip_unreadable:WRITING:{path}:{exc}") from exc
    for token in ("88 lessons", "440 stable required bodies", "0 blockers"):
        if token not in readme:
            raise BaselineError(f"ketw_readme_token_missing:{token}")
    _, bodies = _sheet_rows(body_bytes, "AssetBodies", "asset_id")
    _, bundles = _sheet_rows(body_bytes, "LessonBundles", "lesson_id")
    _require(len(bodies), 440, "count:WRITING:asset_bodies")
    _require(len(bundles), 88, "count:WRITING:lessons")
    _require(len({row.get("asset_id") for row in bodies}), 440, "identity:WRITING:asset_id")
    _require(len({row.get("lesson_id") for row in bundles}), 88, "identity:WRITING:lesson_id")
    _require({row.get("role") for row in bodies}, {"CTX", "NTC", "PRD", "CHK", "EVD"}, "roles:WRITING")
    if any(row.get("disposition") != "ADMITTED_INTERNAL_D0" for row in bodies):
        raise BaselineError("writing_non_admitted_body")
    if any(row.get("maturity") != "CONTENT_REVIEWED_INTERNAL_D0" for row in bodies):
        raise BaselineError("writing_maturity_drift")
    if any(row.get("bundle_status") != "COMPLETE_FIVE_ROLE_BUNDLE" for row in bundles):
        raise BaselineError("writing_bundle_incomplete")
    return {
        "skill": "WRITING",
        "source_filename": path.name,
        "source_sha256": _sha256(path),
        "zip_entry_count": len(names),
        "package_id": "English_Grammar_A1_KETW_Asset_Body_Teacher_Delivery_Internal_D0_v1.0.0",
        "version": "1.0.0",
        "status": "FROZEN_INTERNAL_D0",
        "gate": "PASS_ACCEPTED_AND_FROZEN_INTERNAL_D0",
        "counts": {"lessons": 88, "asset_bodies": 440, "teacher_guides": 88},
        "release_boundary": {
            "internal_ready": 88,
            "learner_release_ready": 0,
            "audio_deferred": False,
        },
    }


def build_baseline(*, ketw: Path, ketr: Path, kets: Path, ketl: Path, source_commit_sha: str) -> dict[str, Any]:
    paths = {"WRITING": ketw, "READING": ketr, "SPEAKING": kets, "LISTENING": ketl}
    for skill, path in paths.items():
        if not path.is_file():
            raise BaselineError(f"package_missing:{skill}:{path}")
    if len(source_commit_sha) != 40 or any(c not in "0123456789abcdef" for c in source_commit_sha.lower()):
        raise BaselineError("source_commit_sha_invalid")
    packages = [
        _writing_package(ketw),
        _manifest_package(ketr, "READING"),
        _manifest_package(kets, "SPEAKING"),
        _manifest_package(ketl, "LISTENING"),
    ]
    totals = {
        key: sum(int(package["counts"][key]) for package in packages)
        for key in ("lessons", "asset_bodies", "teacher_guides")
    }
    return {
        "task_id": TASK_ID,
        "schema_version": SCHEMA_VERSION,
        "validation_status": STATUS,
        "source_commit_sha": source_commit_sha.lower(),
        "packages": packages,
        "totals": totals,
        "claim_boundaries": {
            "metadata_only": True,
            "source_packages_committed": False,
            "asset_body_content_modified": False,
            "learner_release_approved": False,
            "human_pilot_claimed": False,
            "listening_audio_complete": False,
            "a2_unlocked": False,
            "mastery_claimed": False,
        },
        "errors": [],
        "next_short_step": NEXT_SHORT_STEP,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ketw", type=Path, required=True)
    parser.add_argument("--ketr", type=Path, required=True)
    parser.add_argument("--kets", type=Path, required=True)
    parser.add_argument("--ketl", type=Path, required=True)
    parser.add_argument("--source-commit-sha", required=True)
    parser.add_argument("--output-root", type=Path, required=True)
    args = parser.parse_args()
    baseline = build_baseline(
        ketw=args.ketw, ketr=args.ketr, kets=args.kets, ketl=args.ketl,
        source_commit_sha=args.source_commit_sha,
    )
    output = args.output_root / "four_skill_asset_body_baseline.private.json"
    _atomic_json(output, baseline)
    print(json.dumps(baseline, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
