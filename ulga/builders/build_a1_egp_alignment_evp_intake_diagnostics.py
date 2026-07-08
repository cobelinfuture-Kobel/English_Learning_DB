import json
import os
from pathlib import Path

BASE = Path(__file__).resolve().parents[2]
OUT = BASE / "ulga" / "reports" / "a1_egp_alignment_evp_intake_diagnostics.json"
SUMMARY = BASE / "ulga" / "reports" / "a1_egp_alignment_evp_intake_diagnostics_summary.json"
TASK_ID = "R7-M104C_A1EGPAlignmentEVPIntakeDiagnostics"
CANDIDATES = [
    "English Vocabulary Profile Online.xlsx",
    "vocabulary_profile/source/English Vocabulary Profile Online.xlsx",
    "grammar_profile/source/English Vocabulary Profile Online.xlsx",
    "sources/English Vocabulary Profile Online.xlsx",
    "data/source/English Vocabulary Profile Online.xlsx",
]


def write(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def inspect_xlsx(path):
    result = {
        "path": str(path),
        "exists": path.is_file(),
        "readable": False,
        "sheet_names": [],
        "non_empty_preview_rows": 0,
        "error": None,
    }
    if not path.is_file():
        return result
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(path, read_only=True, data_only=True)
        result["sheet_names"] = list(workbook.sheetnames)
        preview_rows = 0
        for sheet_name in workbook.sheetnames[:2]:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=1, max_row=50):
                if any(cell.value is not None for cell in row):
                    preview_rows += 1
        result["non_empty_preview_rows"] = preview_rows
        result["readable"] = preview_rows > 0
    except Exception as exc:
        result["error"] = str(exc)
    return result


def main():
    env_value = os.environ.get("EVP_SOURCE_XLSX")
    inspected = []
    if env_value:
        inspected.append({"source": "EVP_SOURCE_XLSX", **inspect_xlsx(Path(env_value))})
    else:
        inspected.append({"source": "EVP_SOURCE_XLSX", "path": None, "exists": False, "readable": False, "sheet_names": [], "non_empty_preview_rows": 0, "error": "ENV_NOT_SET"})
    for candidate in CANDIDATES:
        inspected.append({"source": "candidate", **inspect_xlsx(BASE / candidate)})
    ready = [item for item in inspected if item.get("readable")]
    recommended = None
    if ready:
        recommended = ready[0]["path"]
    report = {
        "task_id": TASK_ID,
        "artifact_id": "a1_egp_alignment_evp_intake_diagnostics",
        "evp_ready": bool(ready),
        "recommended_evp_source_path": recommended,
        "inspected_paths": inspected,
        "required_operator_commands_if_not_ready": [
            "Create a local folder: mkdir vocabulary_profile\\source",
            "Copy English Vocabulary Profile Online.xlsx into vocabulary_profile\\source\\English Vocabulary Profile Online.xlsx",
            "Or set $env:EVP_SOURCE_XLSX to the exact existing workbook path before running cross-source triage",
            "Then rerun: python ulga/builders/build_a1_egp_alignment_cross_source_triage.py",
        ],
        "final_closeout_allowed": False,
        "a2_a2plus_progression_allowed": False,
        "canonical_grammar_write_allowed": False,
        "local_validation_required": True,
        "ci_gate_required": False,
    }
    summary = {
        "task_id": TASK_ID,
        "artifact_id": "a1_egp_alignment_evp_intake_diagnostics_summary",
        "validation_status": "PASS",
        "evp_ready": bool(ready),
        "recommended_evp_source_path": recommended,
        "inspected_path_count": len(inspected),
        "existing_path_count": sum(1 for item in inspected if item.get("exists")),
        "readable_path_count": len(ready),
        "source_intake_required": not bool(ready),
        "final_closeout_allowed": False,
        "a2_a2plus_progression_allowed": False,
        "canonical_grammar_write_allowed": False,
        "next_short_step": "R7-M104D_A1EGPAlignmentEVPSourceIntakeOrRerunCrossSourceTriage",
        "stop_reason": "SOURCE_INTAKE_REQUIRED" if not ready else "NONE",
    }
    write(OUT, report)
    write(SUMMARY, summary)
    print("A1 EGP alignment EVP intake diagnostics build: PASS")
    print("EVP ready:", summary["evp_ready"])
    print("Existing paths:", summary["existing_path_count"])
    print("Readable paths:", summary["readable_path_count"])
    if recommended:
        print("Recommended EVP source:", recommended)


if __name__ == "__main__":
    main()
