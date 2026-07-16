#!/usr/bin/env python3
"""Build and query the private four-skill Asset Body consumer index.

Task: A1FS-V1-M2_FourSkillAssetBodyConsumerAndQuery
A2 content remains unreadable through the learning query.  Only metadata-only
A2 handoff inspection is exposed until the runtime lock is implemented in M4.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook

from ulga.builders import build_a1fs_v1_m1_prerequisite_graph_and_coverage as m1

TASK_ID = "A1FS-V1-M2_FourSkillAssetBodyConsumerAndQuery"
SCHEMA_VERSION = "a1fs.v1.m2.four_skill_asset_body_consumer.v1"
STATUS = "PASS_A1FS_V1_M2_FOUR_SKILL_ASSET_BODY_CONSUMER_READY"
NEXT_SHORT_STEP = "A1FS-V1-M3_LearnerProfileSessionAndStateStorage"
GRAPH_STATUS = m1.STATUS
MAX_QUERY_LIMIT = 100


class ConsumerError(ValueError):
    """Fail-closed index or query error."""


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _canonical_digest(value: Any) -> str:
    return _sha256_bytes(json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8"))


def _atomic_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.replace(temporary, path)


def _load_json(path: Path, code: str) -> tuple[dict[str, Any], bytes]:
    try:
        raw = path.read_bytes(); value = json.loads(raw)
    except (OSError, json.JSONDecodeError) as exc:
        raise ConsumerError(f"{code}_unreadable:{exc}") from exc
    if not isinstance(value, dict):
        raise ConsumerError(f"{code}_not_object")
    return value, raw


def _json_assets(path: Path, skill: str, member: str) -> list[dict[str, Any]]:
    try:
        with ZipFile(path) as archive:
            payload = json.loads(archive.read(member).decode("utf-8"))
    except Exception as exc:
        raise ConsumerError(f"asset_registry_unreadable:{skill}:{exc}") from exc
    rows = payload.get("rows") or payload.get("asset_bodies")
    if not isinstance(rows, list) or not rows:
        raise ConsumerError(f"asset_registry_empty:{skill}")
    records = []
    for row in rows:
        lesson_id = str(row.get("lesson_id") or ""); asset_id = str(row.get("asset_body_id") or "")
        role = str(row.get("role") or ""); level = m1._level_for(skill, lesson_id, row.get("level"))
        body = row.get("body")
        if not lesson_id or not asset_id or not role or not isinstance(body, dict):
            raise ConsumerError(f"asset_body_invalid:{skill}:{asset_id}")
        records.append({
            "asset_id": asset_id, "lesson_id": lesson_id, "skill": skill,
            "level": level, "role": role, "payload": body,
            "content_digest": _canonical_digest(body),
            "release_scope": "PRIVATE_INTERNAL_D0",
        })
    return records


def _writing_assets(path: Path) -> list[dict[str, Any]]:
    member = "outputs/ketw_asset_body/KETW_AB_Asset_Body_Per_Asset_Admission_Ledger_v1.0.0.xlsx"
    try:
        with ZipFile(path) as archive:
            raw = archive.read(member)
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        rows = list(workbook["AssetBodies"].iter_rows(values_only=True))
    except Exception as exc:
        raise ConsumerError(f"asset_registry_unreadable:WRITING:{exc}") from exc
    index = next((i for i, row in enumerate(rows) if row and row[0] == "asset_id"), None)
    if index is None:
        raise ConsumerError("writing_asset_header_missing")
    headers = [str(value) if value is not None else "" for value in rows[index]]
    fields = (
        "body_title", "body_text", "teacher_delivery", "scaffold_and_fade",
        "expected_evidence", "acceptance_rule", "critical_failure", "diagnostic_route",
    )
    stage_level = {"A1 WF": "A1", "A1+ WB": "A1+", "A2 KP": "A2"}
    records = []
    for values in rows[index + 1 :]:
        if not values or values[0] is None:
            continue
        row = {headers[i]: value for i, value in enumerate(values) if i < len(headers) and headers[i]}
        asset_id = str(row.get("asset_id") or ""); lesson_id = str(row.get("lesson_id") or "")
        role = str(row.get("role") or ""); stage = str(row.get("stage") or "")
        payload = {field: row.get(field) for field in fields if row.get(field) is not None}
        if not asset_id or not lesson_id or not role or stage not in stage_level or not payload:
            raise ConsumerError(f"asset_body_invalid:WRITING:{asset_id}")
        records.append({
            "asset_id": asset_id, "lesson_id": lesson_id, "skill": "WRITING",
            "level": stage_level[stage], "role": role, "payload": payload,
            "content_digest": _canonical_digest(payload),
            "release_scope": "PRIVATE_INTERNAL_D0",
        })
    if not records:
        raise ConsumerError("asset_registry_empty:WRITING")
    return records


def build_index(*, graph_path: Path, baseline_path: Path, ketw: Path, ketr: Path, kets: Path, ketl: Path) -> dict[str, Any]:
    paths = {"WRITING": ketw, "READING": ketr, "SPEAKING": kets, "LISTENING": ketl}
    m1._baseline(baseline_path, paths)
    graph, graph_raw = _load_json(graph_path, "graph")
    if graph.get("validation_status") != GRAPH_STATUS:
        raise ConsumerError("graph_status_invalid")
    baseline_raw = baseline_path.read_bytes()
    if graph.get("source_baseline_sha256") != _sha256_bytes(baseline_raw):
        raise ConsumerError("graph_baseline_binding_invalid")
    assets = (
        _json_assets(ketl, "LISTENING", "ketl_asset_body_production/data/asset_body_registry.json")
        + _json_assets(kets, "SPEAKING", "07_KETS_AB/data/asset_body_registry.json")
        + _json_assets(ketr, "READING", "ketr_asset_body_production/data/asset_body_registry.json")
        + _writing_assets(ketw)
    )
    asset_ids = [row["asset_id"] for row in assets]
    if len(asset_ids) != len(set(asset_ids)):
        # Asset IDs are only guaranteed skill-local; make the public key explicit.
        for row in assets:
            row["asset_key"] = f'{row["skill"]}:{row["asset_id"]}'
    else:
        for row in assets:
            row["asset_key"] = row["asset_id"]
    if len({row["asset_key"] for row in assets}) != len(assets):
        raise ConsumerError("asset_key_collision")
    graph_lessons = {
        row["source_ref"]: row for row in graph.get("nodes", [])
        if row.get("node_type") == "LESSON"
    }
    by_lesson: dict[str, list[dict[str, Any]]] = {}
    for asset in assets:
        if asset["lesson_id"] not in graph_lessons:
            raise ConsumerError(f'asset_lesson_not_in_graph:{asset["skill"]}:{asset["lesson_id"]}')
        node = graph_lessons[asset["lesson_id"]]
        if node.get("skill") != asset["skill"] or node.get("level") != asset["level"]:
            raise ConsumerError(f'asset_graph_partition_mismatch:{asset["asset_key"]}')
        by_lesson.setdefault(asset["lesson_id"], []).append(asset)
    requirements: dict[str, list[str]] = {}
    for row in graph.get("coverage", []):
        for lesson_id in row.get("lesson_ids", []):
            requirements.setdefault(lesson_id, []).append(row["node_id"])
    catalog = []
    for lesson_id, rows in by_lesson.items():
        first = rows[0]
        catalog.append({
            "lesson_id": lesson_id, "lesson_node_id": f'LESSON:{first["skill"]}:{lesson_id}',
            "skill": first["skill"], "level": first["level"],
            "asset_keys": sorted(row["asset_key"] for row in rows),
            "roles": sorted({row["role"] for row in rows}),
            "requirement_node_ids": sorted(set(requirements.get(lesson_id, []))),
            "release_scope": "PRIVATE_INTERNAL_D0",
        })
    expected_assets = sum(int(row["counts"]["asset_bodies"]) for row in json.loads(baseline_raw)["packages"])
    if len(assets) != expected_assets or len(catalog) != int(graph["counts"]["lesson_count"]):
        raise ConsumerError("consumer_count_mismatch")
    assets.sort(key=lambda row: (row["skill"], row["level"], row["lesson_id"], row["role"], row["asset_key"]))
    catalog.sort(key=lambda row: (row["skill"], row["level"], row["lesson_id"]))
    return {
        "task_id": TASK_ID, "schema_version": SCHEMA_VERSION, "validation_status": STATUS,
        "source_graph_sha256": _sha256_bytes(graph_raw),
        "asset_records": assets, "lesson_catalog": catalog,
        "counts": {
            "asset_record_count": len(assets), "lesson_count": len(catalog),
            "learning_lesson_count": sum(row["level"] in {"A1", "A1+"} for row in catalog),
            "a2_handoff_lesson_count": sum(row["level"] == "A2" for row in catalog),
        },
        "access_contract": {
            "visibility": "PRIVATE_INTERNAL", "learning_query_levels": ["A1", "A1+"],
            "a2_payload_query_allowed": False, "a2_handoff_metadata_allowed": True,
            "max_query_limit": MAX_QUERY_LIMIT,
            "filter_fields": ["skill", "level", "lesson_id", "role", "requirement_node_id"],
        },
        "claim_boundaries": {
            "learner_ui_implemented": False, "learner_state_implemented": False,
            "planner_implemented": False, "mastery_claimed": False,
            "learner_release_approved": False, "a2_unlocked": False,
        },
        "errors": [], "next_short_step": NEXT_SHORT_STEP,
    }


def query_index(index: dict[str, Any], *, skill: str | None = None, level: str | None = None,
                lesson_id: str | None = None, role: str | None = None,
                requirement_node_id: str | None = None, offset: int = 0, limit: int = 50) -> dict[str, Any]:
    if index.get("validation_status") != STATUS:
        raise ConsumerError("index_status_invalid")
    if skill is not None and skill not in {"LISTENING", "SPEAKING", "READING", "WRITING"}:
        raise ConsumerError("query_skill_invalid")
    if level == "A2":
        raise ConsumerError("A2_PAYLOAD_LOCKED_USE_HANDOFF_METADATA")
    if level is not None and level not in {"A1", "A1+"}:
        raise ConsumerError("query_level_invalid")
    if offset < 0 or limit < 1 or limit > MAX_QUERY_LIMIT:
        raise ConsumerError("query_page_invalid")
    catalog = {row["lesson_id"]: row for row in index.get("lesson_catalog", [])}
    if lesson_id is not None and lesson_id not in catalog:
        raise ConsumerError("query_lesson_not_found")
    if lesson_id is not None and catalog[lesson_id]["level"] == "A2":
        raise ConsumerError("A2_PAYLOAD_LOCKED_USE_HANDOFF_METADATA")
    allowed_lessons = None
    if requirement_node_id is not None:
        allowed_lessons = {
            row["lesson_id"] for row in catalog.values()
            if requirement_node_id in row.get("requirement_node_ids", []) and row["level"] in {"A1", "A1+"}
        }
        if not allowed_lessons:
            raise ConsumerError("query_requirement_not_found_or_not_learning_scope")
    rows = []
    for row in index.get("asset_records", []):
        if row["level"] == "A2": continue
        if skill is not None and row["skill"] != skill: continue
        if level is not None and row["level"] != level: continue
        if lesson_id is not None and row["lesson_id"] != lesson_id: continue
        if role is not None and row["role"] != role: continue
        if allowed_lessons is not None and row["lesson_id"] not in allowed_lessons: continue
        rows.append(row)
    return {
        "query_status": "PASS_PRIVATE_INTERNAL_QUERY", "total_match_count": len(rows),
        "offset": offset, "limit": limit, "returned_count": len(rows[offset:offset + limit]),
        "asset_records": rows[offset:offset + limit],
        "a2_payload_included": False, "learner_release_claimed": False,
    }


def a2_handoff_metadata(index: dict[str, Any], *, skill: str | None = None) -> dict[str, Any]:
    if index.get("validation_status") != STATUS:
        raise ConsumerError("index_status_invalid")
    if skill is not None and skill not in {"LISTENING", "SPEAKING", "READING", "WRITING"}:
        raise ConsumerError("query_skill_invalid")
    lessons = [{
        "lesson_id": row["lesson_id"], "lesson_node_id": row["lesson_node_id"],
        "skill": row["skill"], "level": "A2", "roles": row["roles"],
        "asset_count": len(row["asset_keys"]), "payload_exposed": False,
    } for row in index.get("lesson_catalog", []) if row["level"] == "A2" and (skill is None or row["skill"] == skill)]
    return {"query_status": "PASS_A2_HANDOFF_METADATA_ONLY", "lesson_count": len(lessons), "lessons": lessons, "a2_payload_included": False}


def main() -> int:
    parser = argparse.ArgumentParser(); sub = parser.add_subparsers(dest="command", required=True)
    build = sub.add_parser("build")
    for name in ("graph", "baseline", "ketw", "ketr", "kets", "ketl", "output-root"):
        build.add_argument(f"--{name}", type=Path, required=True)
    query = sub.add_parser("query"); query.add_argument("--index", type=Path, required=True)
    query.add_argument("--skill"); query.add_argument("--level"); query.add_argument("--lesson-id"); query.add_argument("--role"); query.add_argument("--requirement-node-id")
    query.add_argument("--offset", type=int, default=0); query.add_argument("--limit", type=int, default=50)
    handoff = sub.add_parser("handoff-metadata"); handoff.add_argument("--index", type=Path, required=True); handoff.add_argument("--skill")
    args = parser.parse_args()
    if args.command == "build":
        result = build_index(graph_path=args.graph, baseline_path=args.baseline, ketw=args.ketw, ketr=args.ketr, kets=args.kets, ketl=args.ketl)
        output = getattr(args, "output_root") / "four_skill_asset_body_consumer.private.json"; _atomic_json(output, result)
        shown = {"validation_status": STATUS, **result["counts"], "next_short_step": NEXT_SHORT_STEP, "output": str(output)}
    else:
        index, _ = _load_json(args.index, "index")
        shown = a2_handoff_metadata(index, skill=args.skill) if args.command == "handoff-metadata" else query_index(
            index, skill=args.skill, level=args.level, lesson_id=args.lesson_id, role=args.role,
            requirement_node_id=args.requirement_node_id, offset=args.offset, limit=args.limit,
        )
    print(json.dumps(shown, ensure_ascii=False, indent=2)); return 0


if __name__ == "__main__":
    raise SystemExit(main())
