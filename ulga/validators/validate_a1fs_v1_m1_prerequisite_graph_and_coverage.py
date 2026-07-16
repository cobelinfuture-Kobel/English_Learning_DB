#!/usr/bin/env python3
"""Independent validator for the A1FS-V1-M1 graph and coverage output."""
from __future__ import annotations

import argparse
import hashlib
import json
import os
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

TASK_ID = "A1FS-V1-M1_A1A1PlusPrerequisiteGraphAndCoverage"
STATUS = "PASS_A1FS_V1_M1_PREREQUISITE_GRAPH_AND_COVERAGE"
BASELINE_STATUS = "PASS_A1FS_V1_M0_FOUR_SKILL_ASSET_BODY_BASELINE_FROZEN"
NEXT_SHORT_STEP = "A1FS-V1-M2_FourSkillAssetBodyConsumerAndQuery"
REGISTRIES = {
    "LISTENING": "ketl_asset_body_production/data/asset_body_registry.json",
    "READING": "ketr_asset_body_production/data/asset_body_registry.json",
    "SPEAKING": "07_KETS_AB/data/asset_body_registry.json",
}
WRITING_LEDGER = "outputs/ketw_asset_body/KETW_AB_Asset_Body_Per_Asset_Admission_Ledger_v1.0.0.xlsx"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _atomic_json(path: Path, value: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.replace(temporary, path)


def _load_object(path: Path, code: str) -> tuple[dict[str, Any], bytes]:
    raw = path.read_bytes()
    value = json.loads(raw)
    if not isinstance(value, dict):
        raise ValueError(f"{code}_not_object")
    return value, raw


def _actual_counts(path: Path, skill: str) -> tuple[int, int]:
    try:
        with ZipFile(path) as archive:
            bad = archive.testzip()
            if bad is not None:
                raise ValueError(f"zip_crc_failure:{skill}:{bad}")
            if skill == "WRITING":
                workbook = load_workbook(BytesIO(archive.read(WRITING_LEDGER)), read_only=True, data_only=True)
                rows = list(workbook["AssetBodies"].iter_rows(values_only=True))
                index = next(i for i, row in enumerate(rows) if row and row[0] == "asset_id")
                assets = [row for row in rows[index + 1 :] if row and row[0] is not None]
                lesson_index = next(i for i, value in enumerate(rows[index]) if value == "lesson_id")
                return len({str(row[lesson_index]) for row in assets}), len(assets)
            payload = json.loads(archive.read(REGISTRIES[skill]).decode("utf-8"))
            assets = payload.get("rows") or payload.get("asset_bodies")
            if not isinstance(assets, list):
                raise ValueError(f"asset_rows_invalid:{skill}")
            return len({str(row.get("lesson_id")) for row in assets}), len(assets)
    except (OSError, BadZipFile, KeyError, json.JSONDecodeError) as exc:
        raise ValueError(f"package_unreadable:{skill}:{exc}") from exc


def _is_acyclic(node_ids: set[str], edges: list[dict[str, Any]]) -> bool:
    incoming = {node: 0 for node in node_ids}; outgoing: dict[str, list[str]] = defaultdict(list)
    for edge in edges:
        source = edge.get("from_node_id"); target = edge.get("to_node_id")
        if source not in node_ids or target not in node_ids:
            return False
        incoming[target] += 1; outgoing[source].append(target)
    queue = [node for node, count in incoming.items() if count == 0]
    visited = 0
    while queue:
        node = queue.pop(); visited += 1
        for target in outgoing[node]:
            incoming[target] -= 1
            if incoming[target] == 0:
                queue.append(target)
    return visited == len(node_ids)


def validate(graph_path: Path, baseline_path: Path, package_paths: dict[str, Path]) -> dict[str, Any]:
    errors: list[str] = []
    try:
        graph, _ = _load_object(graph_path, "graph")
        baseline, baseline_raw = _load_object(baseline_path, "baseline")
    except (OSError, json.JSONDecodeError, ValueError) as exc:
        return {"validation_status": "FAIL_A1FS_V1_M1", "error_count": 1, "errors": [str(exc)]}
    if graph.get("task_id") != TASK_ID: errors.append("task_id_invalid")
    if graph.get("validation_status") != STATUS: errors.append("validation_status_invalid")
    if graph.get("next_short_step") != NEXT_SHORT_STEP: errors.append("next_short_step_invalid")
    if baseline.get("validation_status") != BASELINE_STATUS: errors.append("baseline_status_invalid")
    if graph.get("source_baseline_sha256") != hashlib.sha256(baseline_raw).hexdigest():
        errors.append("baseline_hash_mismatch")

    expected = {
        row.get("skill"): row for row in baseline.get("packages", []) if isinstance(row, dict)
    }
    for skill, path in package_paths.items():
        if not path.is_file():
            errors.append(f"package_missing:{skill}"); continue
        if expected.get(skill, {}).get("source_sha256") != _sha256(path):
            errors.append(f"package_hash_mismatch:{skill}")
        try:
            lessons, assets = _actual_counts(path, skill)
            counts = expected.get(skill, {}).get("counts", {})
            if lessons != counts.get("lessons"): errors.append(f"lesson_count_mismatch:{skill}")
            if assets != counts.get("asset_bodies"): errors.append(f"asset_count_mismatch:{skill}")
        except ValueError as exc:
            errors.append(str(exc))

    nodes = graph.get("nodes"); edges = graph.get("edges"); coverage = graph.get("coverage")
    if not isinstance(nodes, list) or not isinstance(edges, list) or not isinstance(coverage, list):
        errors.append("graph_collections_invalid"); nodes = []; edges = []; coverage = []
    node_ids = [row.get("node_id") for row in nodes if isinstance(row, dict)]
    node_map = {row.get("node_id"): row for row in nodes if isinstance(row, dict)}
    if len(node_ids) != len(nodes) or len(set(node_ids)) != len(node_ids): errors.append("node_identity_invalid")
    if not _is_acyclic(set(node_ids), [row for row in edges if isinstance(row, dict)]):
        errors.append("graph_not_acyclic_or_edge_unresolved")
    edge_keys = [(row.get("from_node_id"), row.get("to_node_id"), row.get("edge_type")) for row in edges if isinstance(row, dict)]
    if len(edge_keys) != len(edges) or len(set(edge_keys)) != len(edge_keys): errors.append("edge_identity_invalid")

    required = sorted(row["node_id"] for row in nodes if isinstance(row, dict) and row.get("mastery_required_before_a2") is True)
    if any(node_map[node].get("level") not in {"A1", "A1+"} for node in required):
        errors.append("non_prerequisite_level_marked_required")
    if any(node_map[node].get("node_type") == "SUPPORT_RESOURCE" for node in required):
        errors.append("support_resource_became_hidden_gate")
    gate = graph.get("a2_lock_contract") or {}; gate_id = gate.get("gate_node_id")
    if gate.get("state") != "LOCKED_BY_DESIGN" or gate.get("runtime_unlock_implemented") is not False:
        errors.append("a2_lock_boundary_invalid")
    if sorted(gate.get("required_mastery_node_ids") or []) != required:
        errors.append("required_mastery_denominator_mismatch")
    incoming = sorted(row.get("from_node_id") for row in edges if isinstance(row, dict) and row.get("to_node_id") == gate_id and row.get("edge_type") == "UNLOCK_REQUIRES")
    if incoming != required: errors.append("a2_gate_incoming_requirements_mismatch")
    handoff = set(gate.get("a2_handoff_lesson_node_ids") or [])
    actual_handoff = {row["node_id"] for row in nodes if isinstance(row, dict) and row.get("node_type") == "LESSON" and row.get("level") == "A2"}
    if handoff != actual_handoff: errors.append("a2_handoff_lesson_set_mismatch")
    if any(row.get("edge_type") == "PRECEDES" and node_map.get(row.get("from_node_id"), {}).get("level") in {"A1", "A1+"} and node_map.get(row.get("to_node_id"), {}).get("level") == "A2" for row in edges if isinstance(row, dict)):
        errors.append("a2_direct_sequence_bypasses_global_lock")

    coverage_map = {row.get("node_id"): row for row in coverage if isinstance(row, dict)}
    if len(coverage_map) != len(coverage): errors.append("coverage_identity_invalid")
    required_refs = [node for node in required if node_map[node].get("node_type") == "CAPABILITY"]
    for node in required_refs:
        row = coverage_map.get(node)
        if not row or row.get("coverage_status") != "COVERED" or not row.get("lesson_ids") or not row.get("asset_body_ids"):
            errors.append(f"required_node_uncovered:{node}")
    for skill in package_paths:
        levels = {row.get("level") for row in nodes if isinstance(row, dict) and row.get("node_type") == "LESSON" and row.get("skill") == skill}
        if levels != {"A1", "A1+", "A2"}: errors.append(f"skill_level_partition_invalid:{skill}")

    boundaries = graph.get("claim_boundaries") or {}
    for key in ("source_packages_committed", "asset_body_content_modified", "learner_release_approved", "mastery_claimed", "a2_unlocked", "runtime_planner_implemented", "human_pilot_claimed", "listening_audio_complete"):
        if boundaries.get(key) is not False: errors.append(f"claim_boundary_invalid:{key}")
    counts = graph.get("counts") or {}
    expected_counts = {
        "node_count": len(nodes), "edge_count": len(edges), "coverage_record_count": len(coverage),
        "lesson_count": sum(1 for row in nodes if isinstance(row, dict) and row.get("node_type") == "LESSON"),
        "required_mastery_node_count": len(required), "a2_handoff_lesson_count": len(actual_handoff),
    }
    for key, value in expected_counts.items():
        if counts.get(key) != value: errors.append(f"reported_count_mismatch:{key}")
    if counts.get("uncovered_required_node_count") != 0: errors.append("reported_uncovered_count_nonzero")
    return {
        "task_id": TASK_ID,
        "validation_status": STATUS if not errors else "FAIL_A1FS_V1_M1_PREREQUISITE_GRAPH_AND_COVERAGE",
        "error_count": len(errors), "errors": errors,
        "checked_package_count": len(package_paths),
        "checked_node_count": len(nodes), "checked_edge_count": len(edges),
        "checked_required_mastery_node_count": len(required),
        "next_short_step": NEXT_SHORT_STEP if not errors else TASK_ID,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--graph", type=Path, required=True)
    parser.add_argument("--baseline", type=Path, required=True)
    parser.add_argument("--ketw", type=Path, required=True)
    parser.add_argument("--ketr", type=Path, required=True)
    parser.add_argument("--kets", type=Path, required=True)
    parser.add_argument("--ketl", type=Path, required=True)
    parser.add_argument("--validation-report", type=Path, required=True)
    args = parser.parse_args()
    report = validate(args.graph, args.baseline, {
        "WRITING": args.ketw, "READING": args.ketr,
        "SPEAKING": args.kets, "LISTENING": args.ketl,
    })
    _atomic_json(args.validation_report, report)
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["error_count"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
